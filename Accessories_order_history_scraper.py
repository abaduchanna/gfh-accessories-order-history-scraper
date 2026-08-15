#!/usr/bin/env python3
"""


GFH Telecom — Accessories Order History Scraper (GUI)
=======================================================
Logs into the CPWH Wireless portal, extracts the accessories order history,
and saves it to Excel. URL/link columns are kept internally and stripped from
the exported Excel file.

Tracking numbers are extracted in two stages:
  1. INLINE scan of each order-history row (no page navigation).
  2. Fallback: fetch each missing invoice PDF IN-MEMORY (no files saved to
     disk) using a parallel thread pool (8x+ faster than sequential fetch)
     and extract the tracking number via pdfplumber.

The GUI exposes a date-range filter so the user can limit which orders are
processed and exported (All / 7d / 30d / 90d / YTD / Custom).
"""

# ── Auto-installer (version-aware) ─────────────────────────────────────────
import subprocess
import sys, subprocess
def _pkg_version(dist):
    try:
        import importlib.metadata as _md
        return _md.version(dist)
    except Exception:
        return None
def _ensure(pip_name, imp_name):
    # In a frozen PyInstaller --onefile EXE, sys.executable IS the EXE itself.
    # Running `exe -m pip install ...` would re-launch the whole app instead of
    # installing, so if any bundled dependency fails to import at startup it
    # recursively spawns more EXE processes -> flooding Task Manager. When
    # frozen, dependencies are already bundled by PyInstaller, so never attempt
    # a pip install / subprocess spawn at all.
    if getattr(sys, "frozen", False):
        return
    if _pkg_version(pip_name) is not None:
        return
    try:
        __import__(imp_name.split(".")[0])
    except ImportError:
        try:
            print(f"Installing {pip_name}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", pip_name, "-q"])
        except Exception as e:
            print(f"  [WARN] could not install {pip_name}: {e}")
for _p, _i in [("selenium", "selenium"), ("pandas", "pandas"), ("openpyxl", "openpyxl"), ("pdfplumber", "pdfplumber"), ("requests", "requests"), ("Pillow", "PIL")]:
    _ensure(_p, _i)

import os
import io
import re
import time
import queue
import threading
import traceback
import zipfile
import urllib.request
from datetime import datetime, date, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import tkinter as tk
from theme_manager import ThemeManager, apply_theme_to_window, get_copyright_year
from header_manager import FixedHeaderManager
from logo_handler import LogoHandler
from tkinter import ttk, scrolledtext, messagebox, filedialog
import base64
import tempfile

# ============================================================================
# BRAND / CONFIGURATION
# ============================================================================
APP_TITLE = "GFH Telecom — Accessories Order History Scraper"

COLOR_NAVY = "#090d26"        # matches theme_manager.py navy — header blends with logo
EMBEDDED_LOGO_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "embedded_logo_b64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "embedded_logo_b64.txt"), "r").read().strip()
EMBEDDED_ICON_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "embedded_icon_b64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "embedded_icon_b64.txt"), "r").read().strip()

COLOR_NAVY_DARK = "#050817"
COLOR_RED = "#E91B2F"
COLOR_RED_DARK = "#b91424"
COLOR_BG = "#f6f7fb"
COLOR_WHITE = "#ffffff"
COLOR_TEXT = "#1c1c1c"
COLOR_LOG_BG = "#0d0d1f"
COLOR_LOG_TEXT = "#d7e3f0"
COPYRIGHT_TEXT = f"Developed by Abad Umair Channa | Copyright © {date.today().year} | All rights reserved."

ICON_ICO_NAME = "gfh_icon.ico"     # used for taskbar + titlebar (Windows .ico)
WORDMARK_PNG_NAME = "GFH_Telecom_Logo.png"     # used in the header (resized at runtime via PIL)
ICON_ICO_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "icon_ico_b64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "icon_ico_b64.txt"), "r").read().strip()

# Configuration (unchanged from the original CLI script)
LOGIN_URL = "https://www.cpwhwireless.com/login"
USERNAME = "nofalgodil@gfhtelecom.com"
PASSWORD = "Verizon123"

downloads_folder = r"C:\Users\AbadUmairChanna\Downloads"
zip_path = os.path.join(downloads_folder, "edgedriver_win64.zip")
driver_path = os.path.join(downloads_folder, "msedgedriver.exe")
download_dir = downloads_folder


def get_script_dir():
    # In a frozen PyInstaller --onefile EXE, bundled data files (header logo,
    # window icon) are extracted to sys._MEIPASS, NOT saved next to the
    # executable. Resolve there first so the header logo actually loads.
    if getattr(sys, "frozen", False):
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass and meipass.strip():
            return meipass
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def _resource_path(name):
    """Resolve a bundled resource from source or PyInstaller _MEIPASS."""
    if getattr(sys, "frozen", False):
        base = getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, name)


# ============================================================================
# EDGE DRIVER — version-aware auto-management
# ============================================================================
def _get_installed_edge_version():
    """Detect installed Microsoft Edge version on Windows."""
    try:
        edge_paths = [
            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
            r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        ]
        for p in edge_paths:
            if os.path.exists(p):
                result = subprocess.run(
                    ["powershell", "-NoProfile", "-Command",
                     f"(Get-Item '{p}').VersionInfo.ProductVersion"],
                    capture_output=True, text=True, timeout=15
                )
                if result.returncode == 0 and result.stdout.strip():
                    return result.stdout.strip()
        # Fallback: Windows registry
        for hive in ["HKEY_CURRENT_USER", "HKEY_LOCAL_MACHINE"]:
            result = subprocess.run(
                ["reg", "query", f"{hive}\\Software\\Microsoft\\Edge\\BLBeacon", "/v", "version"],
                capture_output=True, text=True, timeout=10
            )
            if result.returncode == 0:
                for line in result.stdout.splitlines():
                    if "version" in line.lower():
                        parts = line.split()
                        if parts:
                            return parts[-1]
    except Exception as e:
        print(f"  [WARN] Could not detect Edge version: {e}")
    return None


def _get_cached_driver_version():
    """Return version string of the locally cached msedgedriver.exe, or None."""
    if not os.path.exists(driver_path):
        return None
    try:
        result = subprocess.run(
            [driver_path, "--version"],
            capture_output=True, text=True, timeout=15
        )
        if result.returncode == 0:
            for tok in result.stdout.split():
                if tok and tok[0].isdigit() and "." in tok:
                    return tok
    except Exception:
        pass
    return None


def _download_matching_driver(major_version):
    """Download the Edge WebDriver build matching `major_version` from Microsoft's CDN."""
    if major_version:
        index_url = f"https://msedgedriver.azureedge.net/LATEST_RELEASE_{major_version}_WINDOWS"
        print(f"  Fetching latest driver build for Edge {major_version}.x ...")
    else:
        index_url = "https://msedgedriver.azureedge.net/LATEST_RELEASE"
        print("  Fetching latest stable Edge driver build ...")
    try:
        req = urllib.request.Request(index_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as r:
            full_version = r.read().decode("utf-8").strip()
    except Exception as e:
        print(f"  [ERROR] Could not fetch driver version index: {e}")
        return False

    download_url = f"https://msedgedriver.azureedge.net/{full_version}/edgedriver_win64.zip"
    print(f"  Downloading Edge WebDriver {full_version} ...")
    try:
        for stale in (zip_path, driver_path):
            if os.path.exists(stale):
                try:
                    os.remove(stale)
                except Exception:
                    pass
        urllib.request.urlretrieve(download_url, zip_path)
        print("  Extraction ...")
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(downloads_folder)
        print("  Driver downloaded and extracted.")
        return True
    except Exception as e:
        print(f"  [ERROR] Download/extract failed: {e}")
        return False


def _build_edge_options():
    edge_options = webdriver.EdgeOptions()
    edge_options.add_argument("--disable-gpu")
    edge_options.add_argument("--no-sandbox")
    edge_options.add_experimental_option("excludeSwitches", ["enable-logging"])
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "plugins.always_open_pdf_externally": True,
    }
    edge_options.add_experimental_option("prefs", prefs)
    return edge_options


# ============================================================================
# SCRAPER LOGIC (prints redirected to GUI log)
# ============================================================================
def setup_driver():
    print("Setting up Microsoft Edge driver...")
    os.makedirs(download_dir, exist_ok=True)

    edge_version = _get_installed_edge_version()
    if edge_version:
        print(f"  Installed Edge version: {edge_version}")
    else:
        print("  [WARN] Could not detect installed Edge version; will rely on Selenium Manager.")

    # ---- Strategy 1: Selenium Manager (auto-downloads the matching driver) ----
    try:
        print("  Trying Selenium Manager (auto-download)...")
        driver = webdriver.Edge(options=_build_edge_options())
        driver.maximize_window()
        print("  Edge driver initialized via Selenium Manager.")
        return driver
    except Exception as e:
        print(f"  Selenium Manager failed: {e}")
        print("  Falling back to local cached driver with version check...")

    # ---- Strategy 2: cached driver in Downloads (verify version match) --------
    cached_version = _get_cached_driver_version()
    if cached_version:
        print(f"  Cached driver version: {cached_version}")

    edge_major = edge_version.split(".")[0] if edge_version else None
    cached_major = cached_version.split(".")[0] if cached_version else None

    if edge_major and cached_major and edge_major == cached_major:
        print("  Cached driver matches installed Edge — using it.")
    else:
        if edge_major:
            print(f"  Cached driver major version ({cached_major}) != Edge major version ({edge_major}).")
        print("  Downloading matching driver...")
        if not _download_matching_driver(edge_major):
            print("\n  [ERROR] Could not obtain a matching Microsoft Edge WebDriver.")
            print("  Please download manually from https://developer.microsoft.com/microsoft-edge/tools/webdriver/")
            print(f"  and place msedgedriver.exe at: {driver_path}")
            return None

    if os.path.exists(zip_path) and not os.path.exists(driver_path):
        print("  Extracting cached zip...")
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(downloads_folder)
            print("  Extraction complete.")
        except Exception as e:
            print(f"  [ERROR] Extraction failed: {e}")
            return None

    if not os.path.exists(driver_path):
        print(f"\n  [ERROR] Driver not found at {driver_path}")
        return None

    service = Service(driver_path)
    driver = webdriver.Edge(service=service, options=_build_edge_options())
    driver.maximize_window()
    print("  Edge driver initialized successfully (local cache).")
    return driver


def login(driver, manual_login_confirm):
    """manual_login_confirm: callable, blocks until the user confirms manual login (GUI-safe)."""
    print(f"\nNavigating to {LOGIN_URL}...")
    driver.get(LOGIN_URL)
    time.sleep(3)

    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        username_field = driver.find_element(
            By.CSS_SELECTOR,
            "input[type='email'], input[name='username'], input[id='username'], input[name='email'], input[id='email']"
        )
        username_field.clear()
        username_field.send_keys(USERNAME)
        print("Username entered")

        password_field = driver.find_element(By.CSS_SELECTOR, "input[type='password']")
        password_field.clear()
        password_field.send_keys(PASSWORD)
        print("Password entered")

        login_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "btnLogin"))
        )
        login_button.click()
        print("Login button clicked")

        time.sleep(5)
        print(f"Current URL after login: {driver.current_url}")
        return True

    except Exception as e:
        print(f"Login failed: {e}")
        print("\nAttempting manual login...")
        print("Please log in manually in the browser window.")
        manual_login_confirm()
        return True


def navigate_to_my_account(driver):
    print("\nNavigating to My Account...")
    try:
        my_account_selectors = [
            "a[href='dashboard']",
            "a[href*='dashboard']",
            "//a[contains(@href, 'dashboard')]",
            "//a[contains(text(), 'My Account')]"
        ]
        for selector in my_account_selectors:
            try:
                if selector.startswith("//"):
                    my_account = driver.find_element(By.XPATH, selector)
                else:
                    my_account = driver.find_element(By.CSS_SELECTOR, selector)
                my_account.click()
                print("Clicked My Account link")
                time.sleep(3)
                return True
            except Exception:
                continue
        print("Could not find My Account link automatically")
        return False
    except Exception as e:
        print(f"Error navigating to My Account: {e}")
        return False


# ----------------------------------------------------------------------------
# INLINE TRACKING EXTRACTION (no page navigation — scans the order-history row)
# ----------------------------------------------------------------------------
# Common carrier tracking-number patterns. Tried in priority order.
TRACKING_PATTERNS = [
    re.compile(r"\b(1Z\s*[0-9A-Z]{3}\s*[0-9A-Z]{3}\s*[0-9A-Z]{2}\s*[0-9A-Z]{6}\s*[0-9A-Z]{2})\b"),  # UPS (18 chars total)
    re.compile(r"\b(1Z[0-9A-Z]{16})\b"),               # UPS compact form
    re.compile(r"\b([A-Z]{2}\d{9}[A-Z]{2})\b"),        # DHL Express
    re.compile(r"\b(\d{20,22})\b"),                     # USPS / FedEx long numeric
    re.compile(r"\b(\d{15,16})\b"),                     # USPS shorter numeric
    re.compile(r"\b([A-Z]{3}\d{6,})\b"),                # generic carrier prefix
]

# Carrier keywords that may appear in a tracking link's href.
CARRIER_KEYWORDS = ("fedex", "ups.com", "usps", "dhl", "ontrac", "lasership")


def _extract_tracking_from_row(row):
    """Scan a single order-history <tr> and return the tracking number (or '').

    Strategy (in priority order — first hit wins):
      1. <a> elements whose href points to a carrier tracking page
         → extract the tracking id from the URL or use the link's text.
      2. Any element in the row whose id/class contains 'tracking'
         → use its stripped text, BUT only if the cleaned value matches a
         tracking-number pattern (defensive — avoids false positives on
         unrelated cells).
      3. Any <td>'s visible text matches one of the TRACKING_PATTERNS,
         excluding cells that contain the Order ID itself.
    """
    # ---- Strategy 1: carrier links -------------------------------------------
    try:
        for a in row.find_elements(By.TAG_NAME, "a"):
            href = (a.get_attribute("href") or "").strip()
            if not href:
                continue
            # Carrier keyword check is case-insensitive, but the tracking-number
            # patterns must run against the ORIGINAL-CASE href (carrier tracking
            # IDs are case-sensitive — UPS uses uppercase, etc.).
            href_lower = href.lower()
            if any(k in href_lower for k in CARRIER_KEYWORDS):
                for pat in TRACKING_PATTERNS:
                    m = pat.search(href)
                    if m:
                        return m.group(1).strip()
                # Fallback: the link's visible text if it looks like a tracking id.
                txt = (a.text or "").strip()
                if txt and re.match(r"^[A-Z0-9][A-Z0-9\s-]{7,}$", txt, flags=re.IGNORECASE):
                    return re.sub(r"\s+", "", txt)
    except Exception:
        pass

    # ---- Strategy 2: elements labeled 'tracking' ----------------------------
    try:
        labeled = row.find_elements(
            By.XPATH,
            ".//*[contains(translate(@id, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'tracking') "
            "or contains(translate(@class, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'tracking')]"
        )
        for el in labeled:
            txt = (el.text or "").strip()
            if not txt:
                continue
            # Strip the "Tracking Number:" prefix if present
            cleaned = re.sub(r"^.*?tracking\s*(number)?\s*[:]?\s*", "", txt, flags=re.IGNORECASE).strip()
            if not cleaned or len(cleaned) < 6:
                continue
            # Defensive: only accept the cleaned value if it actually looks
            # like a tracking number (matches one of the carrier patterns).
            for pat in TRACKING_PATTERNS:
                m = pat.search(cleaned)
                if m:
                    return m.group(1).strip()
            # Or, if the cleaned text is short enough to BE a tracking number
            # itself (10-30 chars, alphanumeric), accept it.
            if 8 <= len(cleaned) <= 35 and re.match(r"^[A-Z0-9][A-Z0-9\- ]+$", cleaned, flags=re.IGNORECASE):
                return re.sub(r"\s+", "", cleaned)
    except Exception:
        pass

    # ---- Strategy 3: scan every cell's text for tracking patterns -----------
    try:
        # Grab the Order ID cell value so we can exclude it from matches.
        order_id_text = ""
        try:
            id_cells = row.find_elements(By.TAG_NAME, "td")
            if len(id_cells) >= 2:
                order_id_text = (id_cells[1].text or "").strip()
        except Exception:
            pass

        for cell in row.find_elements(By.TAG_NAME, "td"):
            txt = (cell.text or "").strip()
            if not txt:
                continue
            # Skip cells that are clearly the amount (starts with $) or a date
            if txt.startswith("$") or re.match(r"^\d{1,4}[-/]\d{1,2}[-/]\d{2,4}", txt):
                continue
            # Skip the Order ID cell itself
            if order_id_text and txt == order_id_text:
                continue
            for pat in TRACKING_PATTERNS:
                m = pat.search(txt)
                if m:
                    candidate = m.group(1).strip()
                    if candidate and candidate != order_id_text:
                        return candidate
    except Exception:
        pass

    return ""


# ----------------------------------------------------------------------------
# DATE RANGE FILTERING
# ----------------------------------------------------------------------------
def parse_order_date(date_str):
    """Parse an order date string (MM/DD/YYYY by default). Returns datetime or None."""
    if not date_str:
        return None
    s = date_str.strip()
    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def get_date_range(mode, from_str, to_str):
    """Return (start_datetime, end_datetime) for the given mode.
    For 'all', returns (None, None). For custom, parses from_str and to_str.
    end_date is set to end-of-day (23:59:59)."""
    today = datetime.now().replace(hour=23, minute=59, second=59, microsecond=0)
    if mode == "all":
        return None, None
    if mode == "7d":
        return today - timedelta(days=7), today
    if mode == "30d":
        return today - timedelta(days=30), today
    if mode == "90d":
        return today - timedelta(days=90), today
    if mode == "ytd":
        return datetime(today.year, 1, 1, 0, 0, 0), today
    if mode == "custom":
        start = None
        end = today
        if from_str:
            for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d"):
                try:
                    start = datetime.strptime(from_str.strip(), fmt).replace(hour=0, minute=0, second=0)
                    break
                except ValueError:
                    pass
        if to_str:
            for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d"):
                try:
                    end = datetime.strptime(to_str.strip(), fmt).replace(hour=23, minute=59, second=59)
                    break
                except ValueError:
                    pass
        return start, end
    return None, None


def filter_orders_by_date(orders, start_date, end_date):
    """Return only orders whose Order Date falls within [start_date, end_date].
    If both are None, returns the list unchanged. Orders with unparseable dates
    are kept (so we don't lose data)."""
    if start_date is None and end_date is None:
        return orders
    out = []
    for o in orders:
        d = parse_order_date(o.get("Order Date", ""))
        if d is None:
            out.append(o)
            continue
        if start_date and d < start_date:
            continue
        if end_date and d > end_date:
            continue
        out.append(o)
    return out


# ----------------------------------------------------------------------------
# PDF-BASED TRACKING EXTRACTION (fallback when inline row scan finds nothing)
# ----------------------------------------------------------------------------
# Regex used to find tracking numbers inside the PDF text.
# The PDF format on CPWH Wireless invoices is:
#   "To Tracking No: 1Z0110EY0378657347"
# We match the `1Z` prefix and capture everything after it that's alphanumeric,
# then strip everything before the `1Z` (per user request).
PDF_TRACKING_PATTERN = re.compile(r"\b(1Z[0-9A-Z]{6,})\b")
# Also support other carriers in case the invoice format changes
PDF_TRACKING_PATTERN_GENERIC = re.compile(
    r"\b(1Z[0-9A-Z]{6,}"                        # UPS
    r"|[A-Z]{2}\d{9}[A-Z]{2}"                  # DHL Express
    r"|\d{20,22}"                              # USPS/FedEx long numeric
    r"|\d{15,16})\b"                           # USPS shorter numeric
)


def _extract_tracking_from_pdf(pdf_source):
    """Parse a PDF and return the first tracking number found, or ''.

    `pdf_source` can be:
      - a file path (str)
      - a file-like object / BytesIO (in-memory PDF bytes — no disk file written)

    Strategy:
      1. Extract text from every page using pdfplumber.
      2. Search for the pattern `1Z` followed by 6+ alphanumeric chars.
         (Everything before `1Z` is discarded — per user request.)
      3. If no UPS-style number is found, fall back to other carrier patterns.
    """
    try:
        import pdfplumber
    except ImportError:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", "pdfplumber", "-q"])
            import pdfplumber
        except Exception as e:
            print(f"    [WARN] pdfplumber not available: {e}")
            return ""

    try:
        import io
        # Accept either a path string or a file-like object.
        if isinstance(pdf_source, (str, bytes)) and not isinstance(pdf_source, bytes):
            opener = lambda: pdfplumber.open(pdf_source)
        else:
            opener = lambda: pdfplumber.open(io.BytesIO(pdf_source.read() if hasattr(pdf_source, "read") else pdf_source))

        all_text = ""
        with opener() as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                all_text += page_text + "\n"
    except Exception as e:
        print(f"    [WARN] PDF parse failed: {e}")
        return ""

    if not all_text.strip():
        return ""

    # Priority 1: UPS-style 1Z... tracking numbers (the format used by CPWH).
    # "Remove everything before 1Z" — the regex group already starts at `1Z`.
    ups_matches = PDF_TRACKING_PATTERN.findall(all_text)
    if ups_matches:
        return ups_matches[0].strip()

    # Priority 2: other carrier formats
    generic_matches = PDF_TRACKING_PATTERN_GENERIC.findall(all_text)
    if generic_matches:
        for m in generic_matches:
            if len(m) >= 12:
                return m.strip()

    return ""


def _fetch_one_pdf_tracking(order, session):
    """Worker: fetch ONE invoice PDF in-memory and return (order, tracking, error)."""
    url = order["_invoice_link"]
    order_id = order.get("Order ID", "unknown")
    try:
        resp = session.get(url, timeout=30)
        if resp.status_code != 200:
            return order, "", f"HTTP {resp.status_code}"
        tracking = _extract_tracking_from_pdf(io.BytesIO(resp.content))
        return order, tracking, ""
    except Exception as e:
        return order, "", str(e)


def fetch_tracking_from_invoices(orders_data, ask_yes_no, should_stop, max_workers=8):
    """Fetch invoice PDFs IN PARALLEL using a thread pool, extract tracking numbers.

    - No PDF files are written to disk — all in-memory via requests + pdfplumber.
    - Uses requests.Session with HTTPAdapter for connection pooling (keep-alive).
    - Skips orders that already have a tracking number (set during inline row scan).
    - `max_workers` controls concurrency (default 8 → ~8x speedup over sequential).
    """
    candidates = [
        o for o in orders_data
        if not o.get("Tracking Number")
        and o.get("_invoice_link")
        and o["_invoice_link"].startswith("http")
    ]
    if not candidates:
        already_have = sum(1 for o in orders_data if o.get("Tracking Number"))
        print(f"All {already_have}/{len(orders_data)} orders already have tracking numbers — no PDF fetch needed.")
        return

    est_seconds = max(1, (len(candidates) * 2) // max(1, max_workers))
    do_fetch = ask_yes_no(
        "Fetch Tracking Numbers from Invoices",
        f"{len(candidates)} order(s) need tracking numbers fetched from invoice PDFs.\n\n"
        f"Fetch them in parallel ({max_workers} workers, ~{est_seconds}s estimated)?\n"
        f"No PDF files will be saved to disk."
    )
    if not do_fetch:
        print("Skipped PDF-based tracking fetch.")
        return

    try:
        import requests
    except ImportError:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", "requests", "-q"])
            import requests
        except Exception as e:
            print(f"[ERROR] requests library not available: {e}")
            print("        Cannot fetch tracking numbers from PDFs without it.")
            return

    print(f"\nFetching tracking from {len(candidates)} invoice PDFs in parallel ({max_workers} workers)...")
    fetched = 0
    t0 = time.time()

    session = requests.Session()
    adapter = requests.adapters.HTTPAdapter(
        pool_connections=max_workers, pool_maxsize=max_workers * 2
    )
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update({"User-Agent": "Mozilla/5.0"})

    total = len(candidates)
    completed = 0
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {pool.submit(_fetch_one_pdf_tracking, o, session): o for o in candidates}
        for fut in as_completed(futures):
            if should_stop():
                for f in futures:
                    f.cancel()
                print("  Stop requested — aborting parallel PDF fetch.")
                break
            try:
                order, tracking, err = fut.result()
            except Exception as e:
                order = futures[fut]
                tracking, err = "", str(e)
            completed += 1
            order_id = order.get("Order ID", "?")
            if tracking:
                order["Tracking Number"] = tracking
                fetched += 1
                print(f"  [{completed}/{total}] {order_id}: tracking = {tracking}")
            elif err:
                print(f"  [{completed}/{total}] {order_id}: no tracking ({err})")
            else:
                print(f"  [{completed}/{total}] {order_id}: no tracking in PDF")

    elapsed = time.time() - t0
    total_with_tracking = sum(1 for o in orders_data if o.get("Tracking Number"))
    print(f"\nFetched {fetched}/{len(candidates)} tracking numbers from PDFs in {elapsed:.1f}s")
    print(f"Total orders with tracking: {total_with_tracking}/{len(orders_data)}")


def extract_order_history(driver):
    print("\nExtracting order history...")
    orders_data = []
    try:
        if "dashboard" not in driver.current_url.lower():
            navigate_to_my_account(driver)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-striped, table.table"))
        )
        table = driver.find_element(By.CSS_SELECTOR, "table.table-striped, table.table")
        rows = table.find_elements(By.CSS_SELECTOR, "tbody tr")
        print(f"Found {len(rows)} order rows")

        for idx, row in enumerate(rows, 1):
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) >= 6:
                order_id = cells[1].text.strip()
                order_date = cells[2].text.strip()
                amount = cells[3].text.strip()
                order_status = cells[5].text.strip()

                # Invoice link — internal only (used for the optional invoice download)
                invoice_link = ""
                try:
                    invoice_elem = cells[4].find_element(By.TAG_NAME, "a")
                    invoice_link = invoice_elem.get_attribute("href")
                except Exception:
                    pass

                # View Order link — internal only (kept for future use, hidden from export)
                view_link = ""
                try:
                    view_elem = cells[6].find_element(By.XPATH, ".//a[contains(text(), 'View Order')]")
                    view_link = view_elem.get_attribute("href")
                except Exception:
                    pass

                # Tracking number — extracted INLINE from the row, no navigation.
                tracking_number = _extract_tracking_from_row(row)

                order = {
                    "#": idx,
                    "Order ID": order_id,
                    "Order Date": order_date,
                    "Amount": amount,
                    "Order Status": order_status,
                    "Tracking Number": tracking_number,
                    "_invoice_link": invoice_link,    # internal-only
                    "_view_order_link": view_link,    # internal-only
                }
                orders_data.append(order)
                if tracking_number:
                    print(f"  Extracted order #{idx}: {order_id}  |  {order_date}  |  {amount}  |  {order_status}  |  tracking={tracking_number}")
                else:
                    print(f"  Extracted order #{idx}: {order_id}  |  {order_date}  |  {amount}  |  {order_status}  |  tracking=(none)")

        print(f"\nSuccessfully extracted {len(orders_data)} orders")
        return orders_data

    except Exception as e:
        print(f"Error extracting orders: {e}")
        print(f"Current URL: {driver.current_url}")
        return []


# ============================================================================
# EXCEL EXPORT (professional styling)
# ============================================================================
def save_to_excel(orders_data):
    """Save to Excel with professional styling.
    Internal-only fields (prefixed with `_`) are stripped before export.
    The exported table uses the GFH brand palette: navy header, alternating
    row banding, thin borders, frozen header row, and auto-fit column widths.
    """
    if not orders_data:
        print("\nNo order data to save")
        return None

    # Strip internal-only fields (anything starting with `_`)
    export_rows = []
    for o in orders_data:
        export_rows.append({k: v for k, v in o.items() if not k.startswith("_")})
    df_export = pd.DataFrame(export_rows)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"cpwh_order_history_{timestamp}.xlsx"
    filepath = os.path.join(download_dir, filename)

    # Professional palette (matches the GUI brand colors)
    NAVY_FILL       = PatternFill(start_color="FF161632", end_color="FF161632", fill_type="solid")
    ROW_ALT_FILL    = PatternFill(start_color="FFF2F4F7", end_color="FFF2F4F7", fill_type="solid")
    ROW_WHITE_FILL  = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    HEADER_FONT     = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    BODY_FONT       = Font(name="Calibri", size=10, color="FF1C1C1C")
    ALIGN_CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    THIN_BORDER     = Border(
        left=Side(style="thin", color="FFD0D5DD"),
        right=Side(style="thin", color="FFD0D5DD"),
        top=Side(style="thin", color="FFD0D5DD"),
        bottom=Side(style="thin", color="FFD0D5DD"),
    )

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name='Order History', index=False)
        worksheet = writer.sheets['Order History']

        max_row = worksheet.max_row
        max_col = worksheet.max_column

        # ---- Header row ----
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = NAVY_FILL
            cell.font = HEADER_FONT
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
        worksheet.row_dimensions[1].height = 26

        # ---- Body rows (alternating banding) ----
        center_cols = {"#"}
        for r in range(2, max_row + 1):
            row_fill = ROW_ALT_FILL if (r % 2 == 0) else ROW_WHITE_FILL
            for c in range(1, max_col + 1):
                cell = worksheet.cell(row=r, column=c)
                cell.fill = row_fill
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                header_val = worksheet.cell(row=1, column=c).value
                if header_val in center_cols:
                    cell.alignment = ALIGN_CENTER
                else:
                    cell.alignment = ALIGN_LEFT
            worksheet.row_dimensions[r].height = 18

        # ---- Auto-fit column widths ----
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for r in range(1, max_row + 1):
                val = worksheet.cell(row=r, column=col_idx).value
                if val is None:
                    continue
                length = sum(2 if ord(ch) > 127 else 1 for ch in str(val))
                if length > max_length:
                    max_length = length
            adjusted_width = min(max(max_length + 3, 10), 45)
            worksheet.column_dimensions[col_letter].width = adjusted_width

        # ---- Freeze header + enable autofilter ----
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        # ---- Sheet tab color (brand navy) ----
        worksheet.sheet_properties.tabColor = "161632"

    print(f"\n{'='*60}")
    print("DATA SAVED SUCCESSFULLY!")
    print(f"{'='*60}")
    print(f"File location: {filepath}")
    print(f"Total orders:  {len(orders_data)}")
    print(f"Columns:       {', '.join(df_export.columns)}")
    print(f"{'='*60}")

    print("\nORDER SUMMARY:")
    summary_cols = [c for c in ["Order ID", "Order Date", "Amount", "Order Status", "Tracking Number"] if c in df_export.columns]
    print(df_export[summary_cols].to_string(index=False))

    return filepath


# ============================================================================
# MAIN SCRAPE ROUTINE
# ============================================================================
def run_scrape(log, ask_yes_no, ask_ok, should_stop, date_range=("all", "", ""), max_workers=8):
    """Main scrape routine. `log` replaces print for status; runs on a worker thread.

    Parameters
    ----------
    date_range : tuple (mode, from_str, to_str)
        mode is one of: 'all', '7d', '30d', '90d', 'ytd', 'custom'.
        from_str/to_str are MM/DD/YYYY strings used only when mode='custom'.
    max_workers : int
        Number of parallel workers for the invoice-PDF tracking fetch.
    """
    driver = None
    try:
        driver = setup_driver()
        if not driver:
            log("Driver setup failed — see log above.")
            return

        def manual_login_confirm():
            ask_ok("Manual Login Required",
                   "Please complete login manually in the browser window,\nthen click OK to continue.")

        if login(driver, manual_login_confirm):
            print("\nLogin successful!")
            orders = extract_order_history(driver)

            if orders:
                # ---- Apply date range filter (BEFORE the slow PDF fetch) -------
                mode, from_str, to_str = date_range
                start_date, end_date = get_date_range(mode, from_str, to_str)
                if start_date is not None or end_date is not None:
                    before = len(orders)
                    orders = filter_orders_by_date(orders, start_date, end_date)
                    print(f"\nDate filter ({mode}): {before} -> {len(orders)} orders in range")
                    if not orders:
                        print("No orders fall in the selected date range. Nothing to export.")
                        return
                    # Renumber the filtered list so '#' is sequential
                    for idx, o in enumerate(orders, 1):
                        o["#"] = idx

                # ---- Parallel fetch of tracking numbers from invoice PDFs ------
                # No PDF files saved to disk — all in-memory.
                fetch_tracking_from_invoices(
                    orders, ask_yes_no, should_stop,
                    max_workers=max_workers,
                )

                save_to_excel(orders)
            else:
                print("\nNo orders found in the order history.")
                print("The browser will stay open for you to check manually.")
        else:
            print("\nLogin had issues. Browser will stay open.")

        print("\n" + "="*60)
        print("Browser will remain open for 15 seconds")
        print("="*60)
        for i in range(15, 0, -1):
            if should_stop():
                break
            time.sleep(1)

    except Exception as e:
        print(f"\nError: {e}")
        print(traceback.format_exc())
    finally:
        if driver:
            try:
                driver.quit()
                print("\nBrowser closed.")
            except Exception:
                pass


# ============================================================================
# GUI
# ============================================================================

def _extract_embedded_icon(b64, filename):
    """Decode an embedded base64 icon to a temp file; return path or None."""
    try:
        if not b64:
            return None
        import base64 as _b64, tempfile, os
        target = os.path.join(tempfile.gettempdir(), filename)
        with open(target, "wb") as fh:
            fh.write(_b64.b64decode(b64))
        return target if os.path.isfile(target) else None
    except Exception:
        return None

class _PrintRedirector:
    """Redirects print()/sys.stdout writes into a thread-safe queue for the GUI log."""
    def __init__(self, q):
        self.q = q
    def write(self, text):
        if text and text.strip("\r\n") != "":
            self.q.put(("log", text.rstrip("\n")))
        elif text == "\n":
            pass
    def flush(self):
        pass


class ScraperApp:
    def __init__(self, root):
        self.root = root
        self._wordmark_img = None
        self._stop_flag = False
        self._busy = False
        self._q = queue.Queue()

        root.title(APP_TITLE)
        # Dynamic screen resolution support: size to 90% of the screen and
        # center it (DPI-aware), then stay a normal resizable top-level so
        # Windows Snap (50% left/right, corners, Win+arrow) keeps working.
        self._apply_dynamic_geometry()
        root.configure(bg=COLOR_BG)
        self._set_window_icon(root)

        self.theme_manager = ThemeManager("GFH Accessories Order History Scraper", app_name="gfh-accessories-order-history-scraper")
        self._build_header()
        self._build_body()
        self._build_log_area()
        self._build_copyright_bar()
        apply_theme_to_window(self.root, self.theme_manager)

        self._poll_queue()
        root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ---- window chrome -----------------------------------------------------
    def _apply_dynamic_geometry(self) -> None:
        """Size the window to 90% of the screen and center it.

        Works on any laptop/monitor/PC (1080p, 1440p, 2K, 4K) and respects
        Windows DPI scaling (run after _enable_dpi_awareness()). The window
        stays resizable so Windows Snap gestures keep working — it centers
        on launch, then snaps normally to 50% left/right, corners or via
        Win+arrow shortcuts.
        """
        try:
            root = self.root
            root.update_idletasks()
            sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
            w = max(640, min(int(sw * 0.90), sw - 20))
            h = max(480, min(int(sh * 0.90), sh - 40))
            x = max(0, (sw - w) // 2)
            y = max(0, (sh - h) // 2)
            root.geometry(f"{w}x{h}+{x}+{y}")
            root.minsize(min(820, max(560, sw // 2)),
                         min(640, max(420, sh // 2)))
            root.resizable(True, True)
        except Exception:
            pass


    def _set_window_icon(self, root):
        """Set taskbar + titlebar icon from embedded base64 ICO."""
        import base64, tempfile, atexit, os, sys

        # 1. Try sys._MEIPASS
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            ico_path = os.path.join(meipass, "gfh_icon.ico")
            if os.path.exists(ico_path):
                try:
                    root.iconbitmap(ico_path)
                    root.iconbitmap(ico_path)
                    return
                except Exception:
                    pass

        # 2. Try next to the exe/script
        if getattr(sys, "frozen", False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = get_script_dir()
        ico_path = os.path.join(base_dir, "gfh_icon.ico")
        if os.path.exists(ico_path):
            try:
                root.iconbitmap(ico_path)
                root.iconbitmap(ico_path)
                return
            except Exception:
                pass

        # 3. Decode to %TEMP%
        try:
            data = base64.b64decode(EMBEDDED_ICON_B64.strip())
            tmp_dir = os.environ.get("TEMP", tempfile.gettempdir())
            ico_path = os.path.join(tmp_dir, "gfh_app_icon.ico")
            with open(ico_path, "wb") as f:
                f.write(data)
            root.iconbitmap(ico_path)
            root.iconbitmap(ico_path)
            return
        except Exception:
            pass

        # 4. Last resort
        try:
            data = base64.b64decode(ICON_ICO_B64.strip())
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".ico")
            tmp.write(data)
            tmp.close()
            atexit.register(lambda p=tmp.name: os.path.exists(p) and os.unlink(p))
            root.iconbitmap(tmp.name)
            root.iconbitmap(tmp.name)
            return
        except Exception:
            pass
        icon_dir = get_script_dir()
        ico_path2 = os.path.join(icon_dir, ICON_ICO_NAME)
        try:
            if os.path.exists(ico_path2):
                root.iconbitmap(ico_path2)
                root.iconbitmap(ico_path2)
        except Exception:
            pass

    # ---- UI construction -----------------------------------------------------

    def _extract_embedded(self, b64, filename):
        """Decode an embedded base64 asset into a temp file; return path or None."""
        try:
            if not b64:
                return None
            import base64 as _b64, tempfile, os
            target = os.path.join(tempfile.gettempdir(), filename)
            with open(target, "wb") as fh:
                fh.write(_b64.b64decode(b64))
            return target if os.path.isfile(target) else None
        except Exception:
            return None


    def _lock_header_colors(self, widget, navy):
        """Recursively bind <Enter>/<Leave> on all header widgets to force navy."""
        try:
            widget.bind("<Enter>", lambda e, w=widget, c=navy: w.configure(bg=c) if not isinstance(w, type(None)) else None)
            widget.bind("<Leave>", lambda e, w=widget, c=navy: w.configure(bg=c) if not isinstance(w, type(None)) else None)
        except Exception:
            pass
        try:
            for child in widget.winfo_children():
                self._lock_header_colors(child, navy)
        except Exception:
            pass
    def _build_header(self):
        """Header using FixedHeaderManager with logo."""
        self.header_mgr = FixedHeaderManager(self.root, title="GFH Accessories Order History Scraper")
        self.header_mgr.add_theme_toggle(self.theme_manager, callback=self._apply_theme)
        # FixedHeaderManager now tags ALL its own widgets with _tag="header"
        # in __init__/add_theme_toggle/add_copyright, so no manual tagging needed.
        # Load logo
        _logo_path = _resource_path(WORDMARK_PNG_NAME)
        if os.path.exists(_logo_path):
            self.header_mgr.set_logo(logo_path=_logo_path, text="GFH")


    def _apply_theme(self, colors=None):
        """Apply theme colors to all widgets, not just ttk styles.

        Single source of truth: delegate to theme_manager.apply_theme_to_window(),
        which walks the tree, skips any widget with _tag in PROTECTED_TAGS,
        and handles Frame/Labelframe/Label/Button/Entry/Text/etc.
        """
        if colors is None:
            colors = self.theme_manager.get_colors()
        # theme_manager.apply_theme_to_window handles:
        #   - ttk.Style configuration (clam theme, TFrame/TLabel/TButton/etc.)
        #   - recursive _walk() that skips _tag-protected widgets (header)
        #   - Labelframe (was previously missed → panels stayed white)
        #   - Checkbutton/Radiobutton selectcolor
        self.theme_manager.apply_theme_to_window(self.root)
        # Refresh header toggle button text in case theme changed
        if hasattr(self.header_mgr, 'update_button_text'):
            self.header_mgr.update_button_text()

    def _build_body(self):
        body = tk.Frame(self.root, bg=COLOR_BG)
        body.pack(fill="x", padx=20, pady=(12, 4))

        # ── Top info row: login user + output folder on a single line ────
        info_row = tk.Frame(body, bg=COLOR_BG)
        info_row.pack(fill="x")
        tk.Label(info_row, text=f"Login user: {USERNAME}", bg=COLOR_BG, fg=COLOR_TEXT,
                 font=("Segoe UI", 9)).pack(side="left")
        tk.Label(info_row, text=f"    Excel output: {download_dir}", bg=COLOR_BG, fg="#555",
                 font=("Segoe UI", 9)).pack(side="left")

        # ── Date Range Filter panel (compact: 2 rows) ─────────────────────
        date_frame = tk.LabelFrame(
            body, text="Date Range Filter",
            bg=COLOR_BG, fg=COLOR_TEXT,
            font=("Segoe UI", 9, "bold"),
            padx=10, pady=6,
        )
        date_frame.pack(fill="x", pady=(8, 6))

        self.date_var = tk.StringVar(value="all")
        presets = [
            ("All Dates",    "all"),
            ("7 Days",       "7d"),
            ("30 Days",      "30d"),
            ("90 Days",      "90d"),
            ("This Year",    "ytd"),
            ("Custom",       "custom"),
        ]
        row1 = tk.Frame(date_frame, bg=COLOR_BG)
        row1.pack(fill="x")
        for label, value in presets:
            ttk.Radiobutton(
                row1, text=label, value=value,
                variable=self.date_var,
                command=self._on_date_range_change,
            ).pack(side="left", padx=(0, 8))

        # ── Row 2: From/To entries + parallel workers selector ───────────
        row2 = tk.Frame(date_frame, bg=COLOR_BG)
        row2.pack(fill="x", pady=(4, 0))
        tk.Label(row2, text="From:", bg=COLOR_BG, fg=COLOR_TEXT,
                 font=("Segoe UI", 9)).pack(side="left")
        self.from_entry = tk.Entry(row2, width=12, font=("Segoe UI", 9),
                                    state="disabled")
        self.from_entry.pack(side="left", padx=(4, 10))
        tk.Label(row2, text="To:", bg=COLOR_BG, fg=COLOR_TEXT,
                 font=("Segoe UI", 9)).pack(side="left")
        self.to_entry = tk.Entry(row2, width=12, font=("Segoe UI", 9),
                                  state="disabled")
        self.to_entry.pack(side="left", padx=(4, 10))
        tk.Label(row2, text="(MM/DD/YYYY)",
                 bg=COLOR_BG, fg="#666", font=("Segoe UI", 8)).pack(side="left")

        # Workers selector on the same row, pushed to the right
        tk.Label(row2, text="Workers:", bg=COLOR_BG, fg=COLOR_TEXT,
                 font=("Segoe UI", 9)).pack(side="right")
        self.workers_var = tk.IntVar(value=8)
        ttk.Combobox(row2, textvariable=self.workers_var, width=4,
                      values=[2, 4, 6, 8, 10, 12, 16],
                      state="readonly").pack(side="right", padx=(4, 0))

        s = ttk.Style()
        s.theme_use("clam")
        s.configure("A.TButton", background=COLOR_RED, foreground="#fff",
                    font=("Segoe UI", 10, "bold"), borderwidth=0, padding=(16, 8))
        s.map("A.TButton", background=[("active", COLOR_RED_DARK)])
        s.configure("D.TButton", background=COLOR_NAVY, foreground="#fff",
                    font=("Segoe UI", 10, "bold"), borderwidth=0, padding=(16, 8))
        s.map("D.TButton", background=[("active", COLOR_NAVY_DARK)])

        btn_row = tk.Frame(body, bg=COLOR_BG)
        btn_row.pack(anchor="w", pady=(8, 0))
        self.start_btn = ttk.Button(btn_row, text="▶  Start Scrape", style="A.TButton", command=self._start)
        self.start_btn.pack(side="left")
        self.stop_btn = ttk.Button(btn_row, text="⏹  Stop", style="D.TButton",
                                    command=self._stop, state="disabled")
        self.stop_btn.pack(side="left", padx=8)

    def _on_date_range_change(self):
        """Enable the From/To entry fields only when 'Custom Range' is selected."""
        custom = (self.date_var.get() == "custom")
        new_state = "normal" if custom else "disabled"
        self.from_entry.config(state=new_state)
        self.to_entry.config(state=new_state)

    def _build_log_area(self):
        wrap = tk.Frame(self.root, bg=COLOR_BG)
        wrap.pack(fill="both", expand=True, padx=20, pady=(4, 6))
        self.log_w = scrolledtext.ScrolledText(
            wrap, font=("Consolas", 9), wrap=tk.WORD,
            bg=COLOR_LOG_BG, fg=COLOR_LOG_TEXT, relief="flat"
        )
        self.log_w.pack(fill="both", expand=True)

    def _build_copyright_bar(self):
        bar = tk.Frame(self.root, bg=COLOR_NAVY, height=26)
        bar._tag = "header"
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)
        tk.Label(bar, text=COPYRIGHT_TEXT, bg=COLOR_NAVY, fg="#9d9db8",
                  font=("Segoe UI", 8)).pack(pady=4)

    # ---- log / thread-safe helpers -----------------------------------------
    def _log(self, msg):
        self._q.put(("log", msg))

    def _poll_queue(self):
        try:
            while True:
                item = self._q.get_nowait()
                kind = item[0]
                if kind == "log":
                    self.log_w.insert(tk.END, f"[{datetime.now():%H:%M:%S}] {item[1]}\n")
                    self.log_w.see(tk.END)
                elif kind == "done":
                    self._busy = False
                    self.start_btn.config(state="normal")
                    self.stop_btn.config(state="disabled")
        except queue.Empty:
            pass
        self.root.after(80, self._poll_queue)

    def _ask_yes_no(self, title, message):
        """Called from worker thread — blocks worker until the user answers on the main thread."""
        result_holder = {}
        event = threading.Event()

        def _show():
            result_holder["v"] = messagebox.askyesno(title, message)
            event.set()

        self.root.after(0, _show)
        event.wait()
        return result_holder.get("v", False)

    def _ask_ok(self, title, message):
        event = threading.Event()

        def _show():
            messagebox.showinfo(title, message)
            event.set()

        self.root.after(0, _show)
        event.wait()

    def _should_stop(self):
        return self._stop_flag

    # ---- actions -------------------------------------------------------------
    def _start(self):
        if self._busy:
            return
        self._busy = True
        self._stop_flag = False
        self.start_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.log_w.delete("1.0", tk.END)

        # Gather date-range + worker settings from the GUI
        mode = self.date_var.get()
        if mode == "custom":
            from_str = self.from_entry.get().strip()
            to_str   = self.to_entry.get().strip()
        else:
            from_str, to_str = "", ""
        try:
            workers = int(self.workers_var.get())
        except Exception:
            workers = 8
        workers = max(1, min(workers, 32))

        sys.stdout = _PrintRedirector(self._q)

        def _worker():
            try:
                run_scrape(
                    self._log, self._ask_yes_no, self._ask_ok, self._should_stop,
                    date_range=(mode, from_str, to_str),
                    max_workers=workers,
                )
            finally:
                sys.stdout = sys.__stdout__
                self._q.put(("done",))

        threading.Thread(target=_worker, daemon=True).start()

    def _stop(self):
        self._stop_flag = True
        self._log("Stop requested — will halt after the current step.")

    def _on_close(self):
        self._stop_flag = True
        self.root.destroy()


def _enable_dpi_awareness() -> None:
    """Make Windows report physical pixels so winfo_screen* is accurate on
    high-DPI displays (1080p, 1440p, 2K, 4K, DPI-scaled laptops)."""
    if sys.platform != "win32":
        return
    try:
        import ctypes
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(1)  # system DPI aware
        except Exception:
            ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass


def main():
    _enable_dpi_awareness()
    root = tk.Tk()
    ScraperApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
